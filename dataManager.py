import sqlite3 as sql
from datetime import datetime
from os import mkdir, path, remove, getcwd, linesep
from shutil import rmtree, copyfile
from matplotlib.pyplot import show, plot, savefig, title, xlabel, ylabel, figure, vlines, grid, ylim, legend, xlim, xticks
#import read_historia

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from pandas import DataFrame

nameDB = "data.db"

def resetDatabase():
    conn = sql.connect(nameDB)
    conn.commit()
    try:
        conn.execute(
            """CREATE TABLE CUCHARAS (
                Name text,
                Creation_Date text
            )"""
        )
        conn.commit()
        conn.close()
    except:
        conn.close()
    try:
        if path.isdir("Historial"):
            pass
        else:
            mkdir("Historial")
    except:
            pass

def changeWorkingDirectory(newWorkingDirectory:str):
    try:
        nameTable = "WORKINGDIRECTORY"
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"UPDATE {nameTable} SET Directory = '{newWorkingDirectory}' WHERE Directory='{getWorkingDirectory()}'"
        cursor.execute(instruccion)
        conn.commit()
        conn.close()
        return True
    except:
        return False

def setWorkingDirectory(workingDirectory:str):
    date = datetime.strftime(datetime.now(), '%d/%m/%Y')
    try:
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        conn.execute(
                """CREATE TABLE WORKINGDIRECTORY (
                    Directory text,
                    Creation_Date text
                )"""
            )
        conn.commit()
        instruccion = f"INSERT INTO WORKINGDIRECTORY VALUES ('{workingDirectory}', '{date}')"
        conn.execute(instruccion)
        conn.commit()
        conn.close()
        return True
    except: 
        return False

def getWorkingDirectory():
    try:
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM WORKINGDIRECTORY"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        conn.close()
        return data[0][0]
    except:
        return False

def addCuchara(nameCuchara:str):
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        return False
    else:
        date = datetime.strftime(datetime.now(), '%d/%m/%Y')
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"INSERT INTO CUCHARAS VALUES ('{nameCuchara}', '{date}')"
        conn.execute(instruccion)
        conn.commit()
        nameTable = "CUCHARA_"+nameCuchara
        conn.execute(
                f"""CREATE TABLE {nameTable} (
                    Campanas integer,
                    Creation_Date text,
                    Finish_Date text,
                    Escoria integer
                )"""
            )
        conn.commit()
        conn.close()
        pathDirectory = path.join("Historial", nameTable)
        mkdir(pathDirectory)
        return True

def addCampana(nameCuchara:str, nameCampana:int):
    nameTable = "CUCHARA_"+nameCuchara
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        if isDuplicated(nameTable, "Campanas", nameCampana):
            print("111")
            return False
        else:
            date = datetime.strftime(datetime.now(), '%d/%m/%Y')
            conn = sql.connect(nameDB)
            cursor = conn.cursor()
            instruccion = f"INSERT INTO {nameTable} VALUES ('{nameCampana}', '{date}', '', 0)"
            conn.execute(instruccion)
            conn.commit()
            nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
            conn.execute(
                f"""CREATE TABLE {nameTable} (
                    Colada integer,
                    Creation_Date text,
                    HTmaxCucharaF float,
                    HTmaxCucharaT float,
                    HTmaxZonasF float,
                    HTmaxZonasT float,
                    HTmaxRefF float,
                    HTmaxRefT float,
                    HistoriaF float,
                    HistoriaT text,
                    RiesgoF text,
                    RiesgoT text,
                    observacionF text,
                    observacionT text,
                    MaxA text,
                    MaxC text,
                    RiesgoA text,
                    RiesgoC text,
                    observacionA text,
                    observacionC text
                )"""
            )
            conn.commit()
            conn.close()
            pathDirectory = path.join("Historial/CUCHARA_"+nameCuchara, "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana))
            mkdir(pathDirectory)
            return True
    else:
        print("111")
        return False

def add_Colada(nameCuchara:str, nameCampana:int, nameColada:int, HTmaxCucharaF:float, HTmaxCucharaT:float, HTmaxZonasF:float, HTmaxZonasT:float, HTmaxRefF:float, HTmaxRefT:float, HistoriaF:float, HistoriaT:str, RiesgoF:str, RiesgoT:str, observacionF:str, observacionT:str, maxA:float, maxC:float, RiesgoA:str, RiesgoC:str, observacionA:str, observacionC:str):
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        nameTable = "CUCHARA_"+nameCuchara
        if isDuplicated(nameTable, "Campanas", nameCampana):
            nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
            if isDuplicated(nameTable, "Colada", nameColada):
                return False
            else:
                date = datetime.strftime(datetime.now(), '%d/%m/%Y')
                conn = sql.connect(nameDB)
                cursor = conn.cursor()
                instruccion = f"INSERT INTO {nameTable} VALUES ('{nameColada}', '{date}', '{HTmaxCucharaF}', '{HTmaxCucharaT}', '{HTmaxZonasF}', '{HTmaxZonasT}', '{HTmaxRefF}', '{HTmaxRefT}', '{HistoriaF}', '{HistoriaT}', '{RiesgoF}', '{RiesgoT}', '{observacionF}', '{observacionT}', '{str(maxA)}', '{str(maxC)}', '{str(RiesgoA)}', '{str(RiesgoC)}', '{str(observacionA)}', '{str(observacionC)}')"
                conn.execute(instruccion)
                conn.commit()
                conn.commit()
                conn.close()
                updatePlot(nameCuchara, nameCampana)
                return True
        else:
            return False
    else:
        return False

def searchDB(nameTable:str, nameColumn:str, value):
    data = []
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    try:
        instruccion = f"SELECT * FROM {nameTable} WHERE {nameColumn} = '{value}'"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
    except:
        pass
    conn.close()
    return data

def isDuplicated(nameTable:str, nameColumn:str, value):
    if searchDB(nameTable, nameColumn, value)==[]:
        return False
    else:
        return True

def deleteColada(nameCuchara:str, nameCampana:int, nameColada:int):
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        nameTable = "CUCHARA_"+nameCuchara
        if isDuplicated(nameTable, "Campanas", nameCampana):
            nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
            if isDuplicated(nameTable, "Colada", nameColada):
                conn = sql.connect(nameDB)
                cursor = conn.cursor()
                instruccion = f"DELETE FROM {nameTable} WHERE Colada = {nameColada}"
                cursor.execute(instruccion)
                conn.commit()
                conn.close()
                if getEscoria(nameCuchara) >= int(getNameColadas(nameCuchara, nameCampana)[-1]):
                    modifyEscoria(nameCuchara, nameCampana, 0)
                remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/"+str(nameColada)+"F.jpg")
                remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/"+str(nameColada)+"T.jpg")
                remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/"+str(nameColada)+"A.jpg")
                remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/"+str(nameColada)+"C.jpg")
                if getNameColadas(nameCuchara, nameCampana)[-1] == 0:
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaSupF.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaMedF.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaInfF.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaSupT.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaMedT.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaInfT.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaCaraA.xlsx")
                    remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaCaraC.xlsx")
                else:
                    colada = getNameColadas(nameCuchara, nameCampana)[-1]
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaSupF.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaMedF.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaInfF.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaSupT.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaMedT.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaInfT.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaMedT.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaCaraA.xlsx")
                    deleteExcelStillColada(colada, "Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/HistoriaCaraC.xlsx")
                updatePlot(nameCuchara, str(nameCampana))
                return True
            else:
                return False
        else:
            return False
    else:
        return False

def deleteExcelStillColada(colada:int, path:str):
    workbook = load_workbook(path)
    workbook = workbook.active
    counteri = 0
    newXLSX = []
    for i in workbook.iter_rows(values_only=True):
        if counteri < colada:
            newXLSX.append(i)
        else: pass
        counteri += 1
    df = DataFrame(newXLSX, index=None)
    remove(path)
    df.to_excel(path, index=None, header=None)
    return True

def deleteCampana(nameCuchara:str, nameCampana:int):
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        nameTable = "CUCHARA_"+nameCuchara
        if isDuplicated(nameTable, "Campanas", nameCampana):
            conn = sql.connect(nameDB)
            cursor = conn.cursor()
            instruccion = f"DELETE FROM {nameTable} WHERE Campanas = {nameCampana}"
            cursor.execute(instruccion)
            conn.commit()
            nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
            instruccion = f"DROP TABLE {nameTable}"
            cursor.execute(instruccion)
            conn.commit()
            conn.close()
            rmtree("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana))
            return True
        else:
            return False
    else:
        return False

def deleteCuchara(nameCuchara:str):
    if isDuplicated("CUCHARAS", "Name", nameCuchara):
        if CucharaIsWithSomething(nameCuchara)==True:
            return False
        else:
            conn = sql.connect(nameDB)
            cursor = conn.cursor()
            instruccion = f"DELETE FROM CUCHARAS WHERE Name = '{nameCuchara}'"
            cursor.execute(instruccion)
            conn.commit()
            nameTable = "CUCHARA_"+nameCuchara
            instruccion = f"DROP TABLE {nameTable}"
            cursor.execute(instruccion)
            conn.commit()
            conn.close()
            rmtree("Historial/CUCHARA_"+nameCuchara)
            return True
    else:
        return False

def CucharaIsWithSomething(nameCuchara:str):
    try:
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        nameTable = "CUCHARA_"+nameCuchara
        instruccion = f"SELECT * FROM {nameTable}"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        conn.close()
    except:
        return "No existe esa cuchara"
    if data == []:
        return False
    else:
        return True

def countCucharas():
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM CUCHARAS;"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    return len(data)

def countCampanas(nameCuchara:str):
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    nameTable = "CUCHARA_"+nameCuchara
    instruccion = f"SELECT * FROM {nameTable};"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    return len(data)

def getNameCucharas():
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM CUCHARAS ORDER BY Name"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    for i in range(len(data)):
        data[i] = data[i][0]
    return data

def getNameCampanas(nameCuchara):
    nameTable = "CUCHARA_"+nameCuchara
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM {nameTable} ORDER BY Campanas"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    for i in range(len(data)):
        data[i] = data[i][0]
    return data

def getNameColadas(nameCuchara, nameCampana):
    nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+nameCampana
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    if data == []:
        data = [0, 0]
    else:
        for i in range(len(data)):
            data[i] = data[i][0]
    return data

def getEscoria(nameCuchara:str) -> int:
    nameTable = "CUCHARA_"+nameCuchara
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM {nameTable} ORDER BY Campanas"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    return data[-1][-1]

def getHistoricosCampanaFT(nameCuchara:str, nameCampana:int):
    nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    conn.close()
    colF = []
    HTmaxCucharaF = []
    HTmaxCucharaT = []
    HTmaxZonasF = []
    HTmaxZonasT = []
    HTmaxRefF = []
    HTmaxRefT = []
    for i in range(len(data)):
        colF.append(data[i][0])
        aux = data[i][2]
        HTmaxCucharaF.append(float(aux))
        aux = data[i][3]
        HTmaxCucharaT.append(float(aux))
        aux = data[i][4]
        aux = fixList(aux)
        HTmaxZonasF.append(aux)
        aux = data[i][5]
        aux = fixList(aux)
        HTmaxZonasT.append(aux)
        aux = data[i][6]
        aux = fixList(aux)
        HTmaxRefF.append(aux)
        aux = data[i][7]
        aux = fixList(aux)
        HTmaxRefT.append(aux)
    colT = colF
    return [colF, colT, HTmaxCucharaF, HTmaxCucharaT, HTmaxZonasF, HTmaxZonasT, HTmaxRefF, HTmaxRefT]

def fixList(aux:str):
    lista = aux[1:len(aux)-1]
    lista = [float(value) for value in lista.split(', ')]
    return lista

def modifyEscoria(nameCuchara:str, nameCampana:str, numEscoria:int):
    nameTable = "CUCHARA_"+nameCuchara
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"UPDATE {nameTable} SET Escoria = '{numEscoria}' WHERE Campanas={nameCampana}"
    cursor.execute(instruccion)
    conn.commit()
    conn.close()
    return True

def haveEscoria(nameCuchara:str, nameCampana:str):
    nameTable = "CUCHARA_"+nameCuchara
    conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"SELECT * FROM {nameTable} WHERE Campanas = {nameCampana}"
    cursor.execute(instruccion)
    data = cursor.fetchall()
    conn.commit()
    return data[3]

def getZonas(nameCuchara:str, nameCampana:str)->float:
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY HTmaxCucharaF"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        maxZonasF = fixList(data[-1][10])
        maxZonasT = fixList(data[-1][11])
        return [maxZonasF, maxZonasT]
    except:
        return False

def getMaxHistory(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        Col = []
        maxF = []
        maxT = []
        maxA = []
        maxC = []
        for i in range(len(data)):
            maxF.append(data[i][2])
            maxT.append(data[i][3])
            maxA.append(float(data[i][14]))
            maxC.append(float(data[i][15]))
            Col.append(data[i][0])
        nameTable = "CUCHARA_"+nameCuchara
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Campanas"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        escoria = data[-1][-1]
        return [Col, maxF, maxT, maxA, maxC, escoria]
    except:
        return False

def getZonasEF(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        ColZona = []
        zonasF = []
        zonasT = []
        for i in range(len(data)):
            zonasF.append(data[i][4])
            zonasT.append(data[i][5])
            ColZona.append(data[i][0])
        for i in range(len(zonasF)):
            zonasF[i] = fixList(zonasF[i])
            zonasT[i] = fixList(zonasT[i])
        return [zonasF, zonasT]
    except:
        return False

def getRiesgoEF(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        ColZona = []
        zonasF = []
        zonasT = []
        for i in range(len(data)):
            zonasF.append(data[i][10])
            zonasT.append(data[i][11])
        for i in range(len(zonasF)):
            zonasF[i] = fixList(zonasF[i])
            zonasT[i] = fixList(zonasT[i])
        return [zonasF[-1], zonasT[-1], data[i][16], data[i][17]]
    except:
        return False

def getZonasHistory(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        ColZona = []
        zonasF = []
        zonasT = []
        zonasA = []
        zonasC = []
        for i in range(len(data)):
            zonasF.append(data[i][4])
            zonasT.append(data[i][5])
            zonasA.append(data[i][14])
            zonasC.append(data[i][15])
            ColZona.append(data[i][0])
        HTZonaF1 = []
        HTZonaF2 = []
        HTZonaF3 = []
        HTZonaT1 = []
        HTZonaT2 = []
        HTZonaT3 = []
        for i in range(len(zonasF)):
            zonasF[i] = fixList(zonasF[i])
            HTZonaF1.append(zonasF[i][0])
            HTZonaF2.append(zonasF[i][1])
            HTZonaF3.append(zonasF[i][2])
            zonasT[i] = fixList(zonasT[i])
            HTZonaT1.append(zonasT[i][0])
            HTZonaT2.append(zonasT[i][1])
            HTZonaT3.append(zonasT[i][2])
        return [ColZona, HTZonaF1, HTZonaF2, HTZonaF3, HTZonaT1, HTZonaT2, HTZonaT3, zonasA, zonasC]
    except:
        return False
    
def updatePlot(nameCuchara:str, nameCampana:str):
    [Col, maxF, maxT, maxA, maxC, escoria] = getMaxHistory(nameCuchara, nameCampana)
    [ColZona, HTZonaF1, HTZonaF2, HTZonaF3, HTZonaT1, HTZonaT2, HTZonaT3, zonasA, zonasC] = getZonasHistory(nameCuchara, nameCampana)
    # Elimino Graficos
    try:
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasFrontal.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasTrasero.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasCaraA.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasCaraC.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona1Frontal.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona2Frontal.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona3Frontal.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona1Trasera.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona2Trasera.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZona3Trasera.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZonaCaraA.png")
        remove("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisZonaCaraC.png")
    except:
        pass
    # Frontal
    figure(1)
    figure(1).clear()
    grid()
    if max(maxF)<500:
        if min(maxF)>250:
            ylim (250 , 500)
        else:
            ylim (min(maxF)-50 , 500)
    else:
        if min(maxF)>250:
            ylim (250 , max(maxF)+50)
        else:
            ylim (min(maxF)-50 , max(maxF)+50)
        
    xlim(0, max(Col)+10)
    xticks(range(0, max(Col)+10, 10))
    plot(Col, maxF)
    plot(ColZona, HTZonaF1)
    plot(ColZona, HTZonaF2)
    plot(ColZona, HTZonaF3)
    legend(["Máximas Temperaturas", "Zona 1", "Zona 2", "Zona 3"])
    if int(escoria)==0:
        pass
    else:
        vlines(int(escoria), 100, 500, colors='b', linestyles='dashed', label='')
    title("Análisis Frontal / Cuchara "+nameCuchara+" - Campaña "+str(nameCampana))
    xlabel("Número de Colada")
    ylabel("Temperatura máxima alcanzada[ºC]")
    savefig("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasFrontal"+".png")
    
    # Trasera
    figure(2)
    figure(2).clear()
    grid()
    if max(maxT)<500:
        if min(maxT)>250:
            ylim (250 , 500)
        else:
            ylim (min(maxT)-50 , 500)
    else:
        if min(maxT)>250:
            ylim (250 , max(maxT)+50)
        else:
            ylim (min(maxT)-50 , max(maxT)+50)
    xlim(0, max(Col)+10)
    xticks(range(0, max(Col)+10, 10))
    plot(Col, maxT)
    plot(ColZona, HTZonaT1)
    plot(ColZona, HTZonaT2)
    plot(ColZona, HTZonaT3)
    legend(["Máximas Temperaturas", "Zona 1", "Zona 2", "Zona 3"])
    if int(escoria)==0:
        pass
    else:
        vlines(int(escoria), 100, 500, colors='b', linestyles='dashed', label='')
    title("Análisis Trasero / Cuchara "+nameCuchara+" - Campaña "+str(nameCampana))
    xlabel("Número de Colada")
    ylabel("Temperatura máxima alcanzada")
    savefig("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasTrasero"+".png")
    # Cara A
    figure(3)
    figure(3).clear()
    grid()
    if max(maxA)<500:
        if min(maxA)>250:
            ylim (250 , 500)
        else:
            ylim (min(maxA)-50 , 500)
    else:
        if min(maxA)>250:
            ylim (250 , max(maxA)+50)
        else:
            ylim (min(maxA)-50 , max(maxA)+50)
    xlim(0, max(Col)+10)
    xticks(range(0, max(Col)+10, 10))
    plot(Col, maxA)
    legend(["Zona Cara A"])
    if int(escoria)==0:
        pass
    else:
        vlines(int(escoria), 100, 600, colors='b', linestyles='dashed', label='')
    title("Análisis Cara A / Cuchara "+nameCuchara+" - Campaña "+str(nameCampana))
    xlabel("Número de Colada")
    ylabel("Temperatura máxima alcanzada[ºC]")
    savefig("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasCaraA"+".png")
    # Cara C
    figure(4)
    figure(4).clear()
    grid()
    if max(maxC)<500:
        if min(maxC)>250:
            ylim (250 , 500)
        else:
            ylim (min(maxC)-50 , 500)
    else:
        if min(maxC)>250:
            ylim (250 , max(maxC)+50)
        else:
            ylim (min(maxC)-50 , max(maxC)+50)
    xlim(0, max(Col)+10)
    xticks(range(0, max(Col)+10, 10))
    plot(Col, maxC)
    legend(["Zona Cara C"])
    if int(escoria)==0:
        pass
    else:
        vlines(int(escoria), 100, 600, colors='b', linestyles='dashed', label='')
    title("Análisis Cara C / Cuchara "+nameCuchara+" - Campaña "+str(nameCampana))
    xlabel("Número de Colada")
    ylabel("Temperatura máxima alcanzada[ºC]")
    savefig("Historial/CUCHARA_"+nameCuchara+"/CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)+"/AnalisisTemperaturasCaraC"+".png")

def dataIsCorrupted():
    observations = "Observaciones:\n"
    iteration = 0
    if getWorkingDirectory()==False:
        iteration += 1
        observations += str(iteration) + ". - El programa se ha instalado correctamente (Introducir licencia del proveedor)\n"
        sendEmail(observations, "Interfaz Gráfica Cucharas - ABAXFEM. Programa Instalado Correctamente", "", "")
    else:
        if getWorkingDirectory()==getcwd():
            pass
        else:
            iteration += 1
            observations += str(iteration) + ". - El programa ha sido cambiado de directorio. Por favor revertir cambios, caso contrario, el programa dejará de funcionar.\nContacto: Abaxfem.com\n"
    try:
        if path.isdir("Historial"):
            pass
        else:
            iteration += 1
            observations += str(iteration) + ". - Directorio \'Historial\' eliminado recientemente (Se ha creado un nuevo directorio automáticamente)\n"
            mkdir("Historial")
        cucharas = getNameCucharas()
        for i in cucharas:
            if path.isdir("Historial/CUCHARA_"+i):
                pass
            else:
                iteration += 1
                observations += str(iteration) + ". - Directorio \'Historial/CUCHARA_"+i+"\' eliminado recientemente (Se ha creado un nuevo directorio automáticamente)\n"
                mkdir("Historial/CUCHARA_"+i)
            campanas = getNameCampanas(i)
            for j in campanas:
                if path.isdir("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)):
                    coladas = getNameColadas(i, str(j))
                    if coladas == [0, 0]:
                        pass
                    else:
                        if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasFrontal.png") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasTrasero.png"):
                            pass
                        else: 
                            iteration += 1
                            observations += str(iteration) + ". - Gráficos \'Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasFrontal.png\' y/o \'Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasTrasero.png\' eliminados recientemente (Se han actualizado los gráficos automáticamente)\n"
                            updatePlot(i, str(j))
                        
                        if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfF.xlsx") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfT.xlsx") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedF.xlsx") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedT.xlsx") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupF.xlsx") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupT.xlsx"):
                            pass
                        else:
                            iteration += 1
                            observations += str(iteration) + ". - \'Excel HistoriaSupF.xlsx\', \'HistoriaSupT.xlsx\', \'HistoriaMedF.xlsx\', \'HistoriaMedT.xlsx\', \'HistoriaInfF.xlsx\' y/o \'HistoriaInfT.xlsx\' de esta ubicación: "+getWorkingDirectory()+"/"+"Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/ eliminados recientemente (El programa puede que no funcione correctamente, contactar a Abaxfem.com)\n"
                            texto = {"Observacion": ["Este archivo fue eliminado, contactar con Abaxfem"]}
                            dataframe = DataFrame(texto)
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfF.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfF.xlsx")
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfT.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaInfT.xlsx")
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedF.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedF.xlsx")
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedT.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaMedT.xlsx")
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupF.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupF.xlsx")
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupT.xlsx"):
                                pass
                            else:
                                dataframe.to_excel("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/HistoriaSupT.xlsx")

                        for k in coladas:
                            if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg") and path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg"):
                                pass
                            else:
                                iteration += 1
                                observations += str(iteration) + ". - Imágenes \'Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg\' y/o \'Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg\' eliminadas recientemente\n"
                                if path.isfile("ResourcesFolder/Imagenes/Borrado.jpg"):
                                    if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg"):
                                        pass
                                    else:
                                        copyfile("ResourcesFolder/Imagenes/Borrado.jpg", "Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg")
                                    if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg"):
                                        pass
                                    else:
                                        copyfile("ResourcesFolder/Imagenes/Borrado.jpg", "Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg")
                                else:
                                    if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg"):
                                        pass
                                    else:
                                        copyfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasFrontal.png", "Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"F.jpg")
                                    if path.isfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg"):
                                        pass
                                    else:
                                        copyfile("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/AnalisisTemperaturasFrontal.png", "Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"/"+str(k)+"T.jpg")
                else:
                    iteration += 1
                    observations += str(iteration) + ". - Directorio \'Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j)+"\' eliminado recientemente (Se ha creado un nuevo directorio automáticamente)\n"
                    mkdir("Historial/CUCHARA_"+i+"/CUCHARA_"+i+"_CAMPANA_"+str(j))
        if observations == "Observaciones:\n":
            return False
        else:
            sendEmail(observations, "El programa: 'Interfaz Gráfica Cucharas - ABAXFEM', tiene observaciones.", "", "")
            return True
    except:
        return True

def getCreationDateCampana(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        return data[0][1]
    except:
        return False

def sendEmail(observations:str, subject, path_attach, name_attach):
    try:
        email_sender = "cucharasreporte@outlook.com"
        password = "ABAXFEM2024"
        
        email_receivers = []
        with open("envVar.txt") as archivo:
            for linea in archivo:
                if linea[0]=="#" or linea[0].rstrip(' ')=='':
                    pass
                else:
                    email_receivers.append(linea.rstrip('\n'))
        print(email_receivers)

        date = datetime.strftime(datetime.now(), '%d/%m/%Y')
        hour = datetime.now()

        subject = subject+ " (" + str(date) + ", "+str(hour.hour)+":"+str(hour.minute)+":"+str(hour.second)+")"
        
        mensaje = MIMEMultipart()
        mensaje['From'] = email_sender
        mensaje['To'] = ", ".join(email_receivers)
        mensaje['Subject'] = subject

        if path_attach == "":
            mensaje.attach(MIMEText(observations, 'plain'))
        else:
            mensaje.attach(MIMEText(observations, 'plain'))
            archivo_adjunto = open(path_attach, 'rb')
            adjunto_MIME = MIMEBase('application', 'octet-stream')
            adjunto_MIME.set_payload((archivo_adjunto).read())
            encoders.encode_base64(adjunto_MIME)
            adjunto_MIME.add_header('Content-Disposition', "attachment; filename= %s" % name_attach)
            mensaje.attach(adjunto_MIME)
        sesion_smtp = smtplib.SMTP('smtp-mail.outlook.com', 587)
        sesion_smtp.starttls()
        sesion_smtp.login(email_sender,password)
        texto = mensaje.as_string()
        sesion_smtp.sendmail(email_sender, email_receivers, texto)
        sesion_smtp.quit()

        print(observations)
    except:
        pass

def getHistoriaEF(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        HistoriaF = data[-1][8]
        HistoriaT = data[-1][9]
        return [HistoriaF, HistoriaT]
    except:
        return False

def getColadasRiesgos(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        coladas = []
        zonasF = []
        zonasT = []
        for i in range(len(data)):
            coladas.append(data[i][0])
            zonasF.append(data[i][10])
            zonasT.append(data[i][11])
        #print(zonasF)
        for i in range(len(zonasF)):
            zonasF[i] = fixList(zonasF[i])
            zonasT[i] = fixList(zonasT[i])
        FSup = []
        FMed = []
        FInf = []
        TSup = []
        TMed = []
        TInf = []
        for i in range(len(zonasF)):
            FSup.append(zonasF[i][0])
            FMed.append(zonasF[i][1])
            FInf.append(zonasF[i][2])
            TSup.append(zonasT[i][0])
            TMed.append(zonasT[i][1])
            TInf.append(zonasT[i][2])
        return [coladas, FSup, FMed, FInf, TSup, TMed, TInf]
    except Exception as e:
        return print(e)

def getColadasRiesgosAC(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        coladas = []
        CaraA = []
        CaraC = []
        for i in range(len(data)):
            CaraA.append(data[i][17])
            CaraC.append(data[i][18])
        for i in range(len(CaraA)):
            CaraA[i] = fixList(CaraA[i])
            CaraC[i] = fixList(CaraC[i])
        RiesgosCaraA = []
        RiesgosCaraC = []
        for i in CaraA:
            RiesgosCaraA.append(i[0])
        for i in CaraC:    
            RiesgosCaraC.append(i[0])
        return [RiesgosCaraA, RiesgosCaraC]
    except:
        return False

def GetReporteObservaciones(nameCuchara:str, nameCampana:str):
    try:
        nameTable = "CUCHARA_"+nameCuchara+"_CAMPANA_"+str(nameCampana)
        conn = sql.connect(nameDB)
        cursor = conn.cursor()
        instruccion = f"SELECT * FROM {nameTable} ORDER BY Colada"
        cursor.execute(instruccion)
        data = cursor.fetchall()
        conn.commit()
        obs = ""
        for i in range(len(data)):
            obs = obs + str(data[i][0])+". - " + str(data[i][-1]) + "<br/>"
        return obs
    except:
        return False


if __name__ == "__main__":
    #path = "Historial/CUCHARA_1/CUCHARA_1_CAMPANA_9/HistoriaSupF.xlsx"
    #colada = 1
    #deleteExcelStillColada(colada, path)
    #resetDatabase()
    #print(getHistoriaEF("1", "2")[1])
    #print(addCuchara("3"))
    #print(addCampana("3", 10))
    #print(searchDB("CUCHARA_2_CAMPANA_20", "Colada", "20"))
    #print(deleteColada("3", "1", 15))
    #print(deleteCampana("2", 22))
    #print(deleteCuchara("Cucha4"))
    #print(CucharaIsWithSomething("2"))
    #print(countCucharas())
    #print(getNameCucharas())
    #print(getNameCampanas("1")[-1])
    #print(getNameColadas("2","15"))
    #print(countCampanas("2"))
    #print(type(getNameCampanas(getNameCucharas()[0])[2]))
    #print(getEscoria('1'))
    #print(add_Colada("1", 1, 1, [356.086], [300.227], [330.503, 356.086, 355.247],[287.276, 300.227, 300.169], [322.488, 342], [274.001, 282.185], "1", "2", "3", "4", "5", "[[Hola], [Bebe], [Bebe]]", 123.45, 456.78, "97.8", "87.5", "Malo", "Malisimo"))
    #add_Colada(1, 1, 1, 356.086, 3000000000000000000, [1 2 3],                    [1, 2, 3],                   [refracft   ], [Refactttttttttttt], 1, 2   , 3   , 4 , 5 ,    [                     ], maxA:float, maxC:float, RiesgoA:str, RiesgoC:str, observacionA:str, observacionC:str):
    #print(getHistoricosCampanaFT("2", 1))
    #print(fixList("[12.12, 24.24, 26.26]"))
    #print(modifyEscoria("2", "22", 50))
    #updatePlot("3", "3")
    #[col, maxF, maT] = getMaxHistory("1", "1")
    #print(maxF)
    #print(isDuplicated("CUCHARAS", "Name", 'Pepe5'))
    #print(getNameCucharas())
    #print(setWorkingDirectory())
    #print(getWorkingDirectory())
    #print(changeWorkingDirectory(getcwd()))
    #print(updatePlot("Cucha45", "1"))
    #print(dataIsCorrupted())
    #print(getZonasHistory("1", "1"))
    #print(getZonas("Cucha45", "100"))
    #print(int(getNameColadas("1", "3")[-1]))
    #print(deleteColada("1", "1", 1))
    #print(getZonasEF('1', '9'))
    #print(getRiesgoEF('1', '9'))
    #print(getColadasRiesgos('3', '1'))
    #sendEmail("Prueba 1", "Fin de Cuchara", "test2.pdf", "test2.pdf")
    #print(GetReporteObservaciones("21", "1"))
    #print(modifyEscoria(nameCampana="11", nameCuchara="1", numEscoria=22))
    #resetDatabase()
    #updatePlot("3", "1")
    #print(getColadasRiesgosAC("1", "1"))
    print(getColadasRiesgos("6", "1"))
    pass
    '''conn = sql.connect(nameDB)
    cursor = conn.cursor()
    instruccion = f"DELETE FROM CUCHARAS WHERE Creation_Date = '27/11/2023'"
    cursor.execute(instruccion)
    conn.commit()
    conn.close()'''